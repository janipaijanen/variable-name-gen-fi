#!/usr/bin/env python3
# encoding: utf-8
# (C) Copyright 2018 Jani Päijänen
import argparse
import os, sys
import openpyxl
from openpyxl import load_workbook
import logging
import json


def find_duplicates(wb:openpyxl.workbook, male_sheetname, female_sheetname) -> {} :
    logger = logging.getLogger(__name__)
    # return: list of duplicate names, e.g. can be given either to a girl or a boy
    females = []
    males_dict = {}
    females_dict = {}
    male_dups = []

    male_max = 0
    male_min = None
    male_total = 0

    female_max = 0
    female_min = None
    female_total = 0

    header_done = False

    for row in wb[female_sheetname].iter_rows():
        name = None
        count = 0

        if header_done == False:
            header_done = True
            continue

        for cell in row:
            if name is None:
                name = cell.value
            else:

                count = cell.value
                female_total = female_total + count

                if count > female_max:
                    female_max = count
                    female_total = male_total + count

        females_dict[name] = count

    logger.debug("male max: {}".format(female_max))
    logger.debug("male total: {}".format(female_total))

    dict_male_dups = {}
    males = []

    female_dup_max = 0

    for row in wb[male_sheetname].iter_rows():
        name = None
        count = 0

        neutral_name = False
        column=-1

        for cell in row:
            column += 1
            # Is it gender neutral name?
            if cell.value in females_dict:
                name = cell.value
                neutral_name = True

            if column == 0:
                if neutral_name == False:
                    males.append(cell.value)

                continue


            if name is not None:
                count = cell.value
                female_total = male_total + count

                if count > male_max:
                    male_max = count

                if male_min is None:
                    male_min = count
                elif count < male_min:
                    male_min = count

                if females_dict[name] > female_dup_max:
                    female_dup_max = females_dict[name]


        if name is not None:
            #logger.debug("{}:{}".format(name, count))
            male_dups.append( (name, count)  )

    logger.debug ("male_min: {}".format(male_min))
    logger.debug ("male_max: {}".format(male_max))
    logger.debug ("female_dup_max: {}".format(female_dup_max))

    scales = {}
    scale_max = None
    scale_min = None
    name_max = None
    name_min = None

    for (name, count) in male_dups:
        duh = females_dict[name] / count * 1.0
        scales[name] = duh

        if scale_max is None:
            scale_max = duh
            name_max = name
        elif duh > scale_max:
            scale_max = duh
            name_max = name

        if scale_min is None:
            scale_min = duh
            name_min = name
        elif duh < scale_min:
            scale_min = duh
            name_min = name

    logger.debug("scale_min {:+.6f}, scale_max {:+.6f}".format(scale_min, scale_max))
    logger.debug("name_min {}, name_max {}".format(name_min, name_max))

    # As zero is male value and one is female value,
    # we want to distinquish gender neutral names from gender target names.
    min_target = 0.001
    max_target = 0.99

    for (name, count) in male_dups:

        dict_male_dups[name] = (
                # scales to (0 .. 1) and works ok, but we need something else
                #( scales[name] - scale_min) / (scale_max - scale_min)

                # scales to (0.05 .. 0.95) https://en.wikipedia.org/wiki/Normalization_(statistics)
                min_target + ((scales[name] - scale_min) * (max_target - min_target) ) / (scale_max - scale_min)

            )

    grande_gender = {}

    for name in females_dict:
        grande_gender[name] = 1.0

    for name in males:
        grande_gender[name] = 0.0

    for name in dict_male_dups:
        grande_gender[name] = dict_male_dups[name]

    return grande_gender


def process_gender(filename:str) -> None:

    logger = logging.getLogger(__name__)
    wb = load_workbook(filename, read_only=True)

    ws_malename = "Miehet kaikki"
    ws_femalename = "Naiset kaikki"

    ws = wb[ws_malename]
    ws = wb[ws_femalename]
    ws = None

    # Column A: Given name
    # Column B: Name count
    name_items = find_duplicates(wb, male_sheetname=ws_malename, female_sheetname=ws_femalename)

    import json
    dump = json.dumps(name_items, ensure_ascii=False, sort_keys=True, indent=4)
    license_text = ""
    with open("./source-data/LICENSE.txt") as fh:
        for line in fh.readlines():
            license_text = "#" + line + license_text

    gender_names = license_text + "gender_names = \\ \n" + dump
    print(gender_names)

def _get_givennames(wb:openpyxl.workbook, sheetname:str, has_header:bool) -> {}:
    logger = logging.getLogger(__name__)
    header_done = False
    names = {}

    for row in wb[sheetname].iter_rows():

        if header_done == False and has_header == True:
            header_done = True
            continue

        name = None
        for cell in row:
            if name is None:
                name = cell.value
            else:
                names[name] = cell.value

    return names

def _process_frequency_givenname(filename:str) -> {}:
    logger = logging.getLogger(__name__)
    wb = load_workbook(filename, read_only=True)
    ws_malename = "Miehet kaikki"
    ws_femalename = "Naiset kaikki"

    try:
        ws = wb[ws_malename]
        ws = wb[ws_femalename]
    except Exception as error:
        logger.error("Filename was: {}".format(filename))
        raise

    stats1 = _get_givennames(wb, ws_malename, has_header=True)
    stats2 = _get_givennames(wb, ws_femalename, has_header=True)
    givennames = {**stats1, **stats2}
    total_givennames = len(givennames)
    stats_givename = {}

    # calculate freqs from stats
    for (givenname,count) in givennames.items():
        stats_givename[givenname] = count / total_givennames * 1.0

    logger.debug (stats_givename)

    return stats_givename

def _get_surnames(wb:openpyxl.workbook, sheetname:str, has_header:bool) -> {}:
    logger = logging.getLogger(__name__)

    header_done = False
    names = {}

    for row in wb[sheetname].iter_rows():

        if header_done == False and has_header == True:
            header_done = True
            continue

        name = None
        for cell in row:
            if name is None:
                name = cell.value
            else:
                names[name] = cell.value

    return names

def _process_frequency_surname(filename:str) -> {}:
    logger = logging.getLogger(__name__)
    wb = load_workbook(filename, read_only=True)
    ws_surnames = "Nimet"

    try:
        ws = wb[ws_surnames]
    except Exception as error:
        logger.error("Filename was: {}".format(filename))
        raise

    surnames = _get_surnames(wb, ws_surnames, has_header=True)

    total_surnames = len(surnames)
    stats_surnames = {}

    # calculate freqs from stats
    for (surname, count) in surnames.items():
        stats_surnames[surname] = count / total_surnames * 1.0

    logger.debug (stats_surnames)
    return stats_surnames

def render_python(attributename:str, freq_dict:{}, license_file:str) -> str:
    logger = logging.getLogger(__name__)
    logger.error ("Not yet implemented.")


    dump = json.dumps(freq_dict, ensure_ascii=False, sort_keys=True, indent=4)
    license_text = ""
    with open(license_file) as fh:
        for line in fh.readlines():
            license_text = "#" + line + license_text

    gender_names = license_text + attributename + " = \\ \n" + dump
    print(gender_names)

    return None


def process_frequency(filename_given_names:str, filename_surnames:str) -> None:
    logger = logging.getLogger(__name__)

    # count all given names and frequencies (0..1)
    freq_given_names = _process_frequency_givenname(filename_given_names)

    # count all surnames and frequencies (0..1)
    freq_surnames = _process_frequency_surname(filename_surnames)

    license_file = "./source-data/LICENSE.txt"

    render_python("given_name_freq", freq_given_names, license_file)
    render_python("surname_freq", freq_surnames, license_file)


def main():
    format='%(levelname)s:%(message)s'
    logging.basicConfig(format=format, level=logging.DEBUG)

    parser = argparse.ArgumentParser()
    parser.add_argument("-g", "--given-names", type=str, default=None, help="Given names input file provided by Finnish Population Register Centre")
    parser.add_argument("-s", "--surnames", type=str, default=None, help="Surnames input file provided by Finnish Population Register Centre")
    parser.add_argument("-pg", "--print-gender",  action='store_true', default=False, help="Print gender from file provided by Finnish Population Register Centre")
    parser.add_argument("-pf", "--print-frequency", action='store_true', default=False, help="Print frequency provided by Finnish Population Register Centre")

    args = parser.parse_args()

    if args.print_gender is not False and (args.given_names is None or os.path.isfile(args.given_names) is False):
        parser.print_usage()
        sys.exit()

    if args.print_frequency is not False and (args.surnames is None or os.path.isfile(args.surnames) is False):
        parser.print_usage()
        sys.exit()

    if args.print_gender == True:
        process_gender(args.given_names)

    if args.print_frequency == True:
        process_frequency(args.given_names, args.surnames)

if __name__ == '__main__':
    main()
